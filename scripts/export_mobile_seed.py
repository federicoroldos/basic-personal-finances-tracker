from __future__ import annotations

import argparse
import json
from pathlib import Path

from openpyxl import load_workbook


DEFAULT_CATEGORIES = [
    "Supermercado",
    "Comida",
    "Transporte",
    "Juegos",
    "Estudio",
    "Hogar",
    "Salud",
    "Otros",
]


def read_rows(workbook_path: Path, sheet_name: str) -> list[dict]:
    wb = load_workbook(workbook_path)
    ws = wb[sheet_name]
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []
    headers = rows[0]
    return [
        dict(zip(headers, row))
        for row in rows[1:]
        if any(value is not None for value in row)
    ]


def export_seed(source: Path, output: Path) -> None:
    config_rows = read_rows(source, "config")
    tx_rows = read_rows(source, "transactions")
    config = {row["key"]: row["value"] for row in config_rows}

    transactions = []
    categories = set(DEFAULT_CATEGORIES)

    for row in tx_rows:
        category = str(row.get("category") or "Otros").strip()
        categories.add(category)
        tx_type = "fund" if str(row.get("type") or "").strip() == "fund" else "expense"
        account = "usd" if str(row.get("account") or "").strip() == "usd" else "uyu"

        transactions.append(
            {
                "id": int(row.get("id") or 0),
                "date": str(row.get("date") or ""),
                "description": str(row.get("description") or "").strip(),
                "amount": round(float(row.get("amount") or 0), 2),
                "category": "Ingreso" if tx_type == "fund" else category,
                "type": tx_type,
                "account": account,
            }
        )

    payload = {
        "schemaVersion": 1,
        "profileName": "",
        "balances": {
            "uyu": round(float(config.get("balance_uyu", 0) or 0), 2),
            "usd": round(float(config.get("balance_usd", 0) or 0), 2),
        },
        "categories": sorted(categories),
        "transactions": sorted(
            transactions,
            key=lambda txn: str(txn["date"]),
            reverse=True,
        ),
    }

    output.write_text(json.dumps(payload, indent=2, ensure_ascii=False), encoding="utf-8")


def main() -> None:
    parser = argparse.ArgumentParser(
        description="Exporta finance_data.xlsx a un JSON compatible con ClariFi Mobile."
    )
    parser.add_argument(
        "--source",
        default="finance_data.xlsx",
        help="Ruta del archivo Excel origen.",
    )
    parser.add_argument(
        "--output",
        default="clarifi-mobile-seed.json",
        help="Ruta del JSON de salida.",
    )
    args = parser.parse_args()

    source = Path(args.source).resolve()
    output = Path(args.output).resolve()

    if not source.exists():
        raise SystemExit(f"No encontré el archivo origen: {source}")

    export_seed(source, output)
    print(f"JSON exportado en: {output}")


if __name__ == "__main__":
    main()
