from flask import Flask, jsonify, request, render_template_string
from datetime import datetime, timedelta
from openpyxl import Workbook, load_workbook
from threading import Lock
import os, secrets

app = Flask(__name__)
app.secret_key = os.environ.get('SECRET_KEY', secrets.token_hex(32))
DATA_PATH  = os.environ.get('DATA_PATH', 'finance_data.xlsx')
XLSX_LOCK  = Lock()
CATEGORIES = ['Housing','Food','Transport','Entertainment','Health','Other']
ACCOUNTS   = {
    'krw': {'name': 'Korean Won',    'symbol': '₩',   'decimals': 0},
    'uyu': {'name': 'Uruguayan Peso','symbol': '$U',  'decimals': 2},
    'usd': {'name': 'US Dollar',     'symbol': 'US$', 'decimals': 2},
}

# XLSX storage
SHEETS = {
    'config': ['key', 'value'],
    'transactions': ['id', 'date', 'description', 'amount', 'category', 'type', 'account'],
    'fixed_payments': ['id', 'name', 'amount', 'account', 'category', 'day'],
    'fixed_applied': ['payment_id', 'year_month'],
}

def _headers(ws):
    return [c.value for c in ws[1]]

def _ensure_headers(ws, expected):
    if _headers(ws)[:len(expected)] != expected:
        for idx, name in enumerate(expected, start=1):
            ws.cell(row=1, column=idx, value=name)

def _rows(ws):
    headers = _headers(ws)
    out = []
    for values in ws.iter_rows(min_row=2, values_only=True):
        if not any(v is not None for v in values):
            continue
        out.append({headers[i]: values[i] if i < len(values) else None for i in range(len(headers))})
    return out

def _next_id(ws):
    ids = []
    for row in _rows(ws):
        try:
            ids.append(int(row.get('id') or 0))
        except (TypeError, ValueError):
            pass
    return max(ids, default=0) + 1

def _load_wb():
    return load_workbook(DATA_PATH)

def init_data():
    if os.path.exists(DATA_PATH):
        wb = load_workbook(DATA_PATH)
    else:
        wb = Workbook()
        wb.remove(wb.active)

    for sheet, headers in SHEETS.items():
        ws = wb[sheet] if sheet in wb.sheetnames else wb.create_sheet(sheet)
        _ensure_headers(ws, headers)

    config = wb['config']
    existing = {r.get('key') for r in _rows(config)}
    for acc in ACCOUNTS:
        key = f'balance_{acc}'
        if key not in existing:
            config.append([key, 0])

    wb.save(DATA_PATH)

# ── HELPERS ───────────────────────────────────────────────────────────────────
def get_balances():
    out = {acc: 0.0 for acc in ACCOUNTS}
    with XLSX_LOCK:
        wb = _load_wb()
        for r in _rows(wb['config']):
            key = str(r.get('key') or '')
            if not key.startswith('balance_'):
                continue
            acc = key.replace('balance_', '')
            if acc in out:
                out[acc] = float(r.get('value') or 0)
    return out

def set_balance(acc, val):
    dec = ACCOUNTS[acc]['decimals']
    rounded = round(float(val), dec)
    key = f'balance_{acc}'
    with XLSX_LOCK:
        wb = _load_wb()
        ws = wb['config']
        for row_idx in range(2, ws.max_row + 1):
            if ws.cell(row=row_idx, column=1).value == key:
                ws.cell(row=row_idx, column=2, value=rounded)
                break
        else:
            ws.append([key, rounded])
        wb.save(DATA_PATH)

def round_acc(acc, val):
    dec = ACCOUNTS[acc]['decimals']
    return round(float(val), dec)

def build_summary():
    bals = get_balances()
    with XLSX_LOCK:
        wb = _load_wb()
        txns = _rows(wb['transactions'])
        fixed = _rows(wb['fixed_payments'])
        applied = _rows(wb['fixed_applied'])

    txns.sort(key=lambda t: (str(t.get('date') or ''), int(t.get('id') or 0)), reverse=True)
    fixed.sort(key=lambda f: int(f.get('day') or 0))

    applied_set = {(a['payment_id'], a['year_month']) for a in applied}
    today_str   = datetime.now().strftime('%Y-%m-%d')
    this_month  = today_str[:7]
    today_day   = int(today_str[8:])
    cutoff      = (datetime.now() - timedelta(days=30)).strftime('%Y-%m-%d')

    stats = {acc: {'exp_cat': {}, 'monthly': {}, 'last30': 0.0} for acc in ACCOUNTS}
    for t in txns:
        acc = t.get('account') or 'uyu'
        if acc not in stats: continue
        amt = t['amount'] or 0
        ds  = str(t['date']) if t['date'] else ''
        m   = ds[:7]
        if t['type'] == 'expense':
            cat = t['category'] or 'Other'
            stats[acc]['exp_cat'][cat] = stats[acc]['exp_cat'].get(cat, 0) + amt
            if ds >= cutoff: stats[acc]['last30'] += amt
        if m:
            stats[acc]['monthly'].setdefault(m, {'in': 0, 'out': 0})
            if t['type'] == 'fund': stats[acc]['monthly'][m]['in']  += amt
            else:                   stats[acc]['monthly'][m]['out'] += amt

    for acc in stats:
        months = sorted(stats[acc]['monthly'])[-6:]
        stats[acc]['monthly'] = {m: stats[acc]['monthly'][m] for m in months}

    # annotate fixed with due/applied status
    for f in fixed:
        f['applied_this_month'] = (f['id'], this_month) in applied_set
        f['due_this_month']     = f['day'] <= today_day and not f['applied_this_month']

    return {
        'balances': bals,
        'stats':    stats,
        'recent':   txns[:15],
        'fixed':    fixed,
        'due_count': sum(1 for f in fixed if f['due_this_month']),
        'categories': CATEGORIES,
        'accounts':   {k: {'name': v['name'], 'symbol': v['symbol'], 'decimals': v['decimals']} for k, v in ACCOUNTS.items()},
        'total_txns': len(txns),
    }

# ── ROUTES ────────────────────────────────────────────────────────────────────
@app.route('/')
def index(): return render_template_string(HTML)

@app.route('/api/summary')
def api_summary(): return jsonify(build_summary())

@app.route('/api/transactions')
def api_transactions():
    with XLSX_LOCK:
        wb = _load_wb()
        txns = _rows(wb['transactions'])
    txns.sort(key=lambda t: (str(t.get('date') or ''), int(t.get('id') or 0)), reverse=True)
    return jsonify(txns)

def _add_txn(data, txn_type):
    acc = data.get('account', 'uyu')
    if acc not in ACCOUNTS: return jsonify({'ok': False, 'error': 'unknown account'}), 400
    amt     = abs(float(data.get('amount', 0)))
    rounded = round_acc(acc, amt)
    bals    = get_balances()
    new_bal = bals[acc] + rounded if txn_type == 'fund' else bals[acc] - rounded
    new_bal = round_acc(acc, new_bal)
    set_balance(acc, new_bal)
    with XLSX_LOCK:
        wb = _load_wb()
        ws = wb['transactions']
        ws.append([
            _next_id(ws),
            data.get('date', datetime.now().strftime('%Y-%m-%d')),
            data.get('description', ''),
            rounded,
            data.get('category', 'Other'),
            txn_type,
            acc,
        ])
        wb.save(DATA_PATH)
    return jsonify({'ok': True, 'balance': new_bal})

@app.route('/api/fund',    methods=['POST'])
def add_fund():    return _add_txn(request.json, 'fund')

@app.route('/api/expense', methods=['POST'])
def add_expense(): return _add_txn(request.json, 'expense')

@app.route('/api/transactions/<int:tid>', methods=['DELETE'])
def delete_txn(tid):
    with XLSX_LOCK:
        wb = _load_wb()
        ws = wb['transactions']
        found = None
        for row_idx in range(2, ws.max_row + 1):
            if int(ws.cell(row=row_idx, column=1).value or 0) == tid:
                found = row_idx
                break
        if found is None:
            return jsonify({'ok': False}), 404
        headers = _headers(ws)
        t = {headers[col - 1]: ws.cell(row=found, column=col).value for col in range(1, len(headers) + 1)}
        acc = t.get('account') or 'uyu'
        delta = float(t.get('amount') or 0)
        ws.delete_rows(found, 1)
        wb.save(DATA_PATH)

    bal = get_balances()[acc]
    new_bal = round_acc(acc, bal - delta if t.get('type') == 'fund' else bal + delta)
    set_balance(acc, new_bal)
    return jsonify({'ok': True})

@app.route('/api/balance', methods=['POST'])
def api_set_balance():
    d = request.json
    acc = d.get('account', 'uyu')
    if acc not in ACCOUNTS: return jsonify({'ok': False}), 400
    set_balance(acc, float(d.get('balance', 0)))
    return jsonify({'ok': True})

# fixed payments
@app.route('/api/fixed')
def api_fixed():
    with XLSX_LOCK:
        wb = _load_wb()
        rows = _rows(wb['fixed_payments'])
        applied = _rows(wb['fixed_applied'])
    this_month = datetime.now().strftime('%Y-%m')
    today_day  = datetime.now().day
    applied_set = {(a['payment_id'], a['year_month']) for a in applied}
    result = []
    for r in sorted(rows, key=lambda item: int(item.get('day') or 0)):
        d = dict(r)
        d['applied_this_month'] = (r['id'], this_month) in applied_set
        d['due_this_month']     = r['day'] <= today_day and not d['applied_this_month']
        result.append(d)
    return jsonify(result)

@app.route('/api/fixed', methods=['POST'])
def create_fixed():
    d   = request.json
    acc = d.get('account', 'uyu')
    if acc not in ACCOUNTS: return jsonify({'ok': False, 'error': 'unknown account'}), 400
    day = int(d.get('day', 1))
    if not 1 <= day <= 31: return jsonify({'ok': False, 'error': 'day must be 1–31'}), 400
    amt = round_acc(acc, float(d.get('amount', 0)))
    with XLSX_LOCK:
        wb = _load_wb()
        ws = wb['fixed_payments']
        fid = _next_id(ws)
        ws.append([fid, d.get('name', ''), amt, acc, d.get('category', 'Other'), day])
        wb.save(DATA_PATH)
    return jsonify({'ok': True, 'id': fid})

@app.route('/api/fixed/<int:fid>', methods=['DELETE'])
def delete_fixed(fid):
    with XLSX_LOCK:
        wb = _load_wb()
        fixed_ws = wb['fixed_payments']
        for row_idx in range(fixed_ws.max_row, 1, -1):
            if int(fixed_ws.cell(row=row_idx, column=1).value or 0) == fid:
                fixed_ws.delete_rows(row_idx, 1)

        applied_ws = wb['fixed_applied']
        for row_idx in range(applied_ws.max_row, 1, -1):
            if int(applied_ws.cell(row=row_idx, column=1).value or 0) == fid:
                applied_ws.delete_rows(row_idx, 1)
        wb.save(DATA_PATH)
    return jsonify({'ok': True})

@app.route('/api/fixed/<int:fid>/apply', methods=['POST'])
def apply_fixed(fid):
    this_month = datetime.now().strftime('%Y-%m')
    today_str  = datetime.now().strftime('%Y-%m-%d')
    with XLSX_LOCK:
        wb = _load_wb()
        fixed = _rows(wb['fixed_payments'])
        fp = next((row for row in fixed if int(row.get('id') or 0) == fid), None)
        if not fp:
            return jsonify({'ok': False}), 404

        applied_ws = wb['fixed_applied']
        applied = _rows(applied_ws)
        if any(int(a.get('payment_id') or 0) == fid and a.get('year_month') == this_month for a in applied):
            return jsonify({'ok': False, 'error': 'already applied this month'}), 400
        applied_ws.append([fid, this_month])
        wb.save(DATA_PATH)
    return _add_txn({'amount': fp['amount'], 'description': fp['name'],
                     'account': fp['account'], 'category': fp['category'],
                     'date': today_str}, 'expense')

@app.route('/api/fixed/<int:fid>/undo', methods=['POST'])
def undo_fixed(fid):
    this_month = datetime.now().strftime('%Y-%m')
    with XLSX_LOCK:
        wb = _load_wb()
        fixed = _rows(wb['fixed_payments'])
        fp = next((row for row in fixed if int(row.get('id') or 0) == fid), None)
        if not fp:
            return jsonify({'ok': False}), 404

        txn_ws = wb['transactions']
        txns = _rows(txn_ws)
        matches = [
            t for t in txns
            if t.get('description') == fp.get('name')
            and t.get('account') == fp.get('account')
            and t.get('type') == 'expense'
            and str(t.get('date') or '').startswith(this_month)
        ]
        t = max(matches, key=lambda row: int(row.get('id') or 0), default=None)
        if t:
            for row_idx in range(txn_ws.max_row, 1, -1):
                if int(txn_ws.cell(row=row_idx, column=1).value or 0) == int(t.get('id') or 0):
                    txn_ws.delete_rows(row_idx, 1)
                    break

        applied_ws = wb['fixed_applied']
        for row_idx in range(applied_ws.max_row, 1, -1):
            if int(applied_ws.cell(row=row_idx, column=1).value or 0) == fid and applied_ws.cell(row=row_idx, column=2).value == this_month:
                applied_ws.delete_rows(row_idx, 1)
        wb.save(DATA_PATH)

    if t:
        acc = t.get('account')
        bal = get_balances()[acc]
        set_balance(acc, round_acc(acc, bal + float(t.get('amount') or 0)))
    return jsonify({'ok': True})

# ── HTML ──────────────────────────────────────────────────────────────────────

# HTML
HTML = '<!DOCTYPE html>\n<html lang="en" data-theme="dark" data-lang="en">\n<head>\n<meta charset="UTF-8">\n<meta name="viewport" content="width=device-width,initial-scale=1">\n<title>ClariFi</title>\n<link rel="icon" type="image/svg+xml" href="data:image/svg+xml;base64,PHN2ZyB4bWxucz0iaHR0cDovL3d3dy53My5vcmcvMjAwMC9zdmciIHZpZXdCb3g9IjAgMCA2NCA2NCI+PGRlZnM+PGxpbmVhckdyYWRpZW50IGlkPSJhIiB4MT0iMCUiIHkxPSIwJSIgeDI9IjEwMCUiIHkyPSIxMDAlIj48c3RvcCBvZmZzZXQ9IjAlIiBzdG9wLWNvbG9yPSIjMEE4NEZGIi8+PHN0b3Agb2Zmc2V0PSIxMDAlIiBzdG9wLWNvbG9yPSIjQkY1QUYyIi8+PC9saW5lYXJHcmFkaWVudD48L2RlZnM+PHJlY3Qgd2lkdGg9IjY0IiBoZWlnaHQ9IjY0IiByeD0iMTQiIGZpbGw9InVybCgjYSkiLz48dGV4dCB4PSIzMiIgeT0iNDMiIGZvbnQtc2l6ZT0iMjgiIGZvbnQtd2VpZ2h0PSI3MDAiIGZpbGw9IndoaXRlIiB0ZXh0LWFuY2hvcj0ibWlkZGxlIiBmb250LWZhbWlseT0iSGVsdmV0aWNhIj5DPC90ZXh0Pjwvc3ZnPg==">\n<link rel="preconnect" href="https://fonts.googleapis.com">\n<link href="https://fonts.googleapis.com/css2?family=Inter:wght@300;400;500;600;700&display=swap" rel="stylesheet">\n<script src="https://cdnjs.cloudflare.com/ajax/libs/Chart.js/4.4.1/chart.umd.js"></script>\n<style>\n:root{\n  --blue:#0A84FF;--green:#30D158;--red:#FF453A;--orange:#FF9F0A;--purple:#BF5AF2;--teal:#40CBE0;\n  --ease:.22s cubic-bezier(.4,0,.2,1);\n}\n[data-theme="dark"]{\n  --bg:#0a0a0f;--bg2:#111118;\n  --surf:rgba(26,26,36,.85);--surf2:rgba(36,36,50,.70);\n  --bdr:rgba(255,255,255,.07);--bdr2:rgba(255,255,255,.12);\n  --lbl:#f0f0f8;--lbl2:rgba(240,240,248,.58);--lbl3:rgba(240,240,248,.28);\n  --fill:rgba(120,120,140,.20);--fill2:rgba(120,120,140,.14);\n  --orb1:radial-gradient(ellipse 800px 600px at 0% 0%,rgba(10,132,255,.10) 0%,transparent 60%);\n  --orb2:radial-gradient(ellipse 700px 600px at 100% 100%,rgba(191,90,242,.08) 0%,transparent 60%);\n  --cg:rgba(255,255,255,.05);--ct:rgba(240,240,248,.40);\n  --sh:0 4px 24px rgba(0,0,0,.45),0 0 0 .5px rgba(255,255,255,.07);\n  --due-bg:rgba(255,159,10,.09);--due-bdr:rgba(255,159,10,.25);\n}\n[data-theme="light"]{\n  --bg:#f0f2f8;--bg2:#e8eaf2;\n  --surf:rgba(255,255,255,.88);--surf2:rgba(255,255,255,.68);\n  --bdr:rgba(60,60,80,.07);--bdr2:rgba(60,60,80,.13);\n  --lbl:#0d0d18;--lbl2:rgba(13,13,24,.58);--lbl3:rgba(13,13,24,.30);\n  --fill:rgba(100,100,130,.10);--fill2:rgba(100,100,130,.07);\n  --orb1:radial-gradient(ellipse 800px 600px at 0% 0%,rgba(10,132,255,.07) 0%,transparent 60%);\n  --orb2:radial-gradient(ellipse 700px 600px at 100% 100%,rgba(191,90,242,.05) 0%,transparent 60%);\n  --cg:rgba(60,60,80,.06);--ct:rgba(13,13,24,.42);\n  --sh:0 4px 20px rgba(0,0,0,.07),0 0 0 .5px rgba(0,0,0,.05);\n  --due-bg:rgba(255,159,10,.07);--due-bdr:rgba(255,159,10,.20);\n}\n*,*::before,*::after{box-sizing:border-box;margin:0;padding:0}\nhtml{scroll-behavior:smooth}\nbody{font-family:-apple-system,\'Inter\',sans-serif;font-size:14px;line-height:1.5;\n  color:var(--lbl);background:var(--bg);min-height:100vh;\n  -webkit-font-smoothing:antialiased;transition:background .3s,color .3s}\n.bg-scene{position:fixed;inset:0;z-index:0;pointer-events:none;background:var(--bg)}\n.bg-scene::before{content:\'\';position:absolute;inset:0;background:var(--orb1)}\n.bg-scene::after{content:\'\';position:absolute;inset:0;background:var(--orb2)}\n#app{position:relative;z-index:1;min-height:100vh;display:flex;flex-direction:column}\n\n/* ── NAV ── */\nnav{\n  height:58px;padding:0 28px;\n  display:flex;align-items:center;justify-content:space-between;gap:16px;\n  background:var(--surf);\n  backdrop-filter:blur(40px) saturate(1.8);\n  -webkit-backdrop-filter:blur(40px) saturate(1.8);\n  border-bottom:.5px solid var(--bdr2);\n  position:sticky;top:0;z-index:100;\n  box-shadow:0 1px 0 var(--bdr);\n}\n.nav-brand{\n  font-size:20px;font-weight:700;letter-spacing:-.6px;\n  background:linear-gradient(135deg,var(--blue),var(--purple));\n  -webkit-background-clip:text;-webkit-text-fill-color:transparent;background-clip:text;\n  flex-shrink:0;\n}\n.nav-balances{display:flex;gap:8px;align-items:center}\n.nav-bal{\n  display:flex;align-items:center;gap:7px;\n  padding:6px 12px;border-radius:10px;\n  background:var(--fill);border:none;cursor:pointer;\n  transition:background .2s,transform .15s;\n}\n.nav-bal:hover{background:var(--fill2)}\n.nav-bal:active{transform:scale(.97)}\n.nav-bal-label{font-size:11px;font-weight:600;color:var(--lbl3);text-transform:uppercase;letter-spacing:.06em}\n.nav-bal-val{font-size:13px;font-weight:700;letter-spacing:-.3px}\n.nav-bal.krw .nav-bal-val{color:var(--blue)}\n.nav-bal.uyu .nav-bal-val{color:var(--orange)}\n.nav-bal.usd .nav-bal-val{color:var(--green)}\n.nav-controls{display:flex;gap:6px;align-items:center}\n.ctrl-btn{\n  height:32px;padding:0 12px;border-radius:8px;border:none;\n  background:var(--fill);color:var(--lbl2);cursor:pointer;\n  font-size:12px;font-weight:600;letter-spacing:.04em;\n  transition:background .2s;font-family:inherit;\n  display:flex;align-items:center;gap:5px;\n}\n.ctrl-btn:hover{background:var(--fill2)}\n.lang-btn{color:var(--purple)}\n\n/* ── DUE BANNER ── */\n#due-banner{\n  display:none;padding:10px 28px;\n  background:var(--due-bg);border-bottom:.5px solid var(--due-bdr);\n  align-items:center;justify-content:space-between;gap:12px;\n}\n.due-text{font-size:13px;font-weight:500;color:var(--orange);display:flex;align-items:center;gap:7px}\n\n/* -- SEGMENTED TABS -- */\n.seg-wrap{\n  padding:8px 28px 0;\n  background:var(--surf2);\n  backdrop-filter:blur(20px);-webkit-backdrop-filter:blur(20px);\n  border-bottom:.5px solid var(--bdr);\n}\n.seg{\n  display:inline-flex;align-items:center;gap:2px;\n  background:var(--fill);border-radius:14px;padding:2px;margin-bottom:8px;\n}\n.seg-item{\n  min-width:86px;padding:7px 16px;border-radius:11px;\n  font-size:13px;font-weight:600;color:var(--lbl2);\n  cursor:pointer;border:none;background:transparent;font-family:inherit;\n  transition:all .2s var(--ease);white-space:nowrap;position:relative;\n}\n.seg-item:hover{color:var(--lbl)}\n.seg-item:active{transform:scale(.96)}\n.seg-item.active{\n  background:var(--surf);color:var(--lbl);\n  box-shadow:0 1px 4px rgba(0,0,0,.18),0 0 0 .5px var(--bdr2);\n}\n.tab-badge{\n  display:inline-block;background:var(--orange);color:#fff;\n  border-radius:99px;font-size:10px;font-weight:700;\n  padding:1px 6px;margin-left:5px;vertical-align:middle;\n}\n\n/* ── MAIN LAYOUT ── */\n.main{flex:1;padding:28px;max-width:1280px;margin:0 auto;width:100%}\n.page{display:none}.page.active{display:block;animation:fade .2s var(--ease)}\n@keyframes fade{from{opacity:0;transform:translateY(4px)}to{opacity:1;transform:translateY(0)}}\n\n/* ── OVERVIEW GRID ── */\n.ov-grid{display:grid;grid-template-columns:1fr 1fr 1fr;gap:12px;margin-bottom:20px}\n.ov-charts{display:grid;grid-template-columns:1fr 1fr;gap:16px;margin-bottom:16px}\n.ov-bottom{display:grid;grid-template-columns:2fr 1fr;gap:16px}\n\n/* ── GLASS CARD ── */\n.card{\n  background:var(--surf);\n  backdrop-filter:blur(24px) saturate(1.6);\n  -webkit-backdrop-filter:blur(24px) saturate(1.6);\n  border:.5px solid var(--bdr2);border-radius:16px;\n  box-shadow:var(--sh);padding:20px 22px;\n}\n.card-title{font-size:11px;font-weight:600;color:var(--lbl3);\n  text-transform:uppercase;letter-spacing:.08em;margin-bottom:14px}\n\n/* ── METRIC CARD ── */\n.metric{padding:18px 20px}\n.metric-ey{font-size:11px;font-weight:500;color:var(--lbl3);\n  text-transform:uppercase;letter-spacing:.07em;margin-bottom:7px}\n.metric-val{font-size:28px;font-weight:300;letter-spacing:-.9px;line-height:1;color:var(--lbl)}\n.metric-val .sym{font-size:15px;font-weight:600;vertical-align:super;margin-right:2px;opacity:.72;letter-spacing:0}\n.metric-val .dec{font-size:15px;font-weight:400;opacity:.60}\n.metric-val.v-krw{color:var(--blue)}.metric-val.v-uyu{color:var(--orange)}\n.metric-val.v-usd{color:var(--green)}.metric-val.v-red{color:var(--red)}\n.metric-sub{font-size:11px;color:var(--lbl3);margin-top:6px}\n.acc-switcher{display:flex;gap:6px;margin-bottom:18px}\n.acc-pill{\n  padding:5px 14px;border-radius:99px;font-size:12px;font-weight:600;\n  cursor:pointer;border:none;background:var(--fill);color:var(--lbl2);\n  transition:all .2s var(--ease);font-family:inherit;\n}\n.acc-pill:hover{background:var(--fill2)}\n.acc-pill.krw{background:rgba(10,132,255,.14);color:var(--blue)}\n.acc-pill.uyu{background:rgba(255,159,10,.12);color:var(--orange)}\n.acc-pill.usd{background:rgba(48,209,88,.12);color:var(--green)}\n\n/* ── TABLE ── */\ntable{width:100%;border-collapse:collapse}\nth{text-align:left;font-size:10px;font-weight:600;color:var(--lbl3);\n  text-transform:uppercase;letter-spacing:.07em;padding:6px 10px;\n  border-bottom:.5px solid var(--bdr2)}\ntd{padding:10px 10px;font-size:13px;border-bottom:.5px solid var(--bdr);vertical-align:middle}\ntr:last-child td{border-bottom:none}\ntr:hover td{background:rgba(255,255,255,.02)}\n[data-theme="light"] tr:hover td{background:rgba(0,0,0,.02)}\n.td-mono{font-weight:600;letter-spacing:-.3px}\n\n/* ── CHIPS ── */\n.chip{display:inline-flex;align-items:center;padding:2px 8px;border-radius:99px;font-size:11px;font-weight:600}\n.chip-expense{background:rgba(255,69,58,.12);color:var(--red)}\n.chip-fund{background:rgba(48,209,88,.12);color:var(--green)}\n.chip-krw{background:rgba(10,132,255,.12);color:var(--blue)}\n.chip-uyu{background:rgba(255,159,10,.12);color:var(--orange)}\n.chip-usd{background:rgba(48,209,88,.12);color:var(--green)}\n.chip-ok{background:rgba(48,209,88,.12);color:var(--green)}\n\n/* ── ADD PAGE ── */\n.add-layout{display:grid;grid-template-columns:minmax(0,720px);gap:16px;align-items:start;justify-content:center}\n.type-toggle{display:grid;grid-template-columns:1fr 1fr;background:var(--fill);border-radius:10px;padding:3px;gap:3px;margin-bottom:16px}\n.type-btn{padding:9px;border-radius:8px;text-align:center;font-size:13px;font-weight:600;cursor:pointer;\n  border:none;background:transparent;color:var(--lbl2);transition:all .2s var(--ease);font-family:inherit}\n.type-btn.a-ex{background:var(--surf);color:var(--red);box-shadow:0 1px 6px rgba(0,0,0,.18)}\n.type-btn.a-fund{background:var(--surf);color:var(--green);box-shadow:0 1px 6px rgba(0,0,0,.18)}\n.form-grid{display:grid;grid-template-columns:1fr 1fr;gap:12px;margin-bottom:12px}\n.form-full{grid-column:1/-1}\n.field{display:flex;flex-direction:column;gap:5px}\n.field label{font-size:11px;font-weight:500;color:var(--lbl2)}\ninput,select{\n  border:.5px solid var(--bdr2);border-radius:9px;padding:9px 12px;\n  font-size:14px;width:100%;outline:none;\n  background:var(--fill);color:var(--lbl);font-family:inherit;\n  transition:background .2s,box-shadow .2s;-webkit-appearance:none;appearance:none;\n}\ninput:focus,select:focus{background:var(--fill2);box-shadow:0 0 0 3px rgba(10,132,255,.20);border-color:var(--blue)}\ninput::placeholder{color:var(--lbl3)}\nselect option{background:#1a1a24;color:#f0f0f8}\n[data-theme="light"] select option{background:#fff;color:#0d0d18}\n.amt-wrap{position:relative}\n.amt-sym{position:absolute;left:11px;top:50%;transform:translateY(-50%);\n  font-size:12px;font-weight:700;color:var(--lbl3);pointer-events:none}\n.amt-wrap input{padding-left:34px;font-size:16px;font-weight:300;letter-spacing:-.4px}\n\n/* ── BUTTONS ── */\n.btn{padding:10px 20px;border-radius:9px;font-size:14px;font-weight:600;cursor:pointer;\n  border:none;font-family:inherit;transition:all .2s var(--ease);display:inline-flex;align-items:center;gap:6px}\n.btn:active{transform:scale(.98)}\n.btn-blue{background:var(--blue);color:#fff;box-shadow:0 3px 12px rgba(10,132,255,.30)}\n.btn-blue:hover{filter:brightness(1.08)}\n.btn-green{background:var(--green);color:#fff;box-shadow:0 3px 12px rgba(48,209,88,.25)}\n.btn-green:hover{filter:brightness(1.08)}\n.btn-orange{background:var(--orange);color:#fff;box-shadow:0 3px 12px rgba(255,159,10,.25)}\n.btn-ghost{background:var(--fill);color:var(--lbl2);box-shadow:none}\n.btn-ghost:hover{background:var(--fill2)}\n.btn-sm{padding:6px 13px;font-size:12px;border-radius:7px}\n.btn-full{width:100%;justify-content:center}\n.del-btn{width:26px;height:26px;border-radius:50%;background:rgba(255,69,58,.10);\n  color:var(--red);border:none;cursor:pointer;font-size:14px;\n  display:inline-flex;align-items:center;justify-content:center;transition:all .2s;flex-shrink:0}\n.del-btn:hover{background:rgba(255,69,58,.20)}\n\n/* ── HISTORY ── */\n.filter-bar{display:flex;align-items:center;justify-content:space-between;\n  flex-wrap:wrap;gap:10px;margin-bottom:14px}\n.filter-bar select,.filter-bar input[type="month"]{width:auto;font-size:13px;padding:7px 10px;border-radius:8px}\n\n/* ── FIXED PAYMENTS ── */\n.fp-grid{display:grid;grid-template-columns:1fr 1fr;gap:12px;margin-bottom:16px}\n.fp-row{\n  display:flex;align-items:center;gap:12px;\n  padding:14px 16px;border-radius:12px;\n  background:var(--fill);border:.5px solid var(--bdr);box-shadow:none;\n  transition:border-color .2s,background .2s;\n}\n.fp-row.due{border-color:var(--due-bdr);background:var(--due-bg)}\n.fixed-list-card .fp-grid{grid-template-columns:1fr;margin-bottom:0}\n.fixed-list-card .fp-row + .fp-row{margin-top:10px}\n.fp-badge{width:42px;height:42px;border-radius:10px;flex-shrink:0;\n  display:flex;flex-direction:column;align-items:center;justify-content:center}\n.fp-badge .fp-d{font-size:17px;font-weight:700;line-height:1}\n.fp-badge .fp-mo{font-size:8px;font-weight:600;text-transform:uppercase;opacity:.65}\n.fp-badge.krw{background:rgba(10,132,255,.15);color:var(--blue)}\n.fp-badge.uyu{background:rgba(255,159,10,.14);color:var(--orange)}\n.fp-badge.usd{background:rgba(48,209,88,.13);color:var(--green)}\n.fp-info{flex:1;min-width:0}\n.fp-name{font-size:14px;font-weight:500;white-space:nowrap;overflow:hidden;text-overflow:ellipsis}\n.fp-meta{font-size:11px;color:var(--lbl3);margin-top:2px}\n.fp-right{display:flex;flex-direction:column;align-items:flex-end;gap:6px;flex-shrink:0}\n.fp-amt{font-size:14px;font-weight:700;letter-spacing:-.3px}\n.fp-actions{display:flex;gap:5px;align-items:center}\n.fixed-layout{display:grid;grid-template-columns:minmax(0,760px);gap:16px;align-items:start;justify-content:center}\n.fixed-layout > div:first-child{order:2;display:grid;gap:16px}\n.fixed-layout > .card{order:1;position:static!important}\n\n/* ── SETTINGS ── */\n.settings-layout{display:grid;grid-template-columns:1fr 1fr 1fr;gap:14px}\n.settings-row{display:flex;gap:8px;align-items:center;margin-top:10px}\n.settings-row input{max-width:150px}\n\n/* ── MISC ── */\n.msg{font-size:12px;font-weight:500;min-height:18px;margin-top:8px}\n.msg.ok{color:var(--green)}.msg.err{color:var(--red)}\n.empty{text-align:center;padding:32px;font-size:13px;color:var(--lbl3)}\n.sec-label{font-size:11px;font-weight:600;color:var(--lbl3);\n  text-transform:uppercase;letter-spacing:.08em;margin-bottom:10px}\n.divider{height:.5px;background:var(--bdr);margin:18px 0}\n.mt12{margin-top:12px}.mt16{margin-top:16px}\n.recent-preview{max-height:360px;overflow-y:auto}\n.recent-preview::-webkit-scrollbar{width:4px}\n.recent-preview::-webkit-scrollbar-thumb{background:var(--bdr2);border-radius:99px}\n/* Keep WebView/browser compositing stable during tab switches and async data paints. */\n#app[data-ready="false"]{visibility:hidden}\n*,*::before,*::after{animation:none!important;transition:none!important}\n.page.active{display:block;transform:none!important}\n.surf,nav,.top-tabs,.tab-bar,.card{backdrop-filter:none!important;-webkit-backdrop-filter:none!important}\n</style>\n</head>\n<body>\n<div class="bg-scene"></div>\n<div id="app" data-ready="false">\n\n<!-- NAV -->\n<nav>\n  <div class="nav-brand">ClariFi</div>\n  <div class="nav-balances">\n    <button class="nav-bal krw" onclick="showTab(\'overview\');setAcc(\'krw\')">\n      <span class="nav-bal-label">KRW</span>\n      <span class="nav-bal-val" id="nav-krw">—</span>\n    </button>\n    <button class="nav-bal uyu" onclick="showTab(\'overview\');setAcc(\'uyu\')">\n      <span class="nav-bal-label">UYU</span>\n      <span class="nav-bal-val" id="nav-uyu">—</span>\n    </button>\n    <button class="nav-bal usd" onclick="showTab(\'overview\');setAcc(\'usd\')">\n      <span class="nav-bal-label">USD</span>\n      <span class="nav-bal-val" id="nav-usd">—</span>\n    </button>\n  </div>\n  <div class="nav-controls">\n    <button class="ctrl-btn lang-btn" onclick="toggleLang()" id="lang-btn">🌐 EN</button>\n    <button class="ctrl-btn" onclick="toggleTheme()" id="theme-btn">🌙 Dark</button>\n  </div>\n</nav>\n\n<!-- DUE BANNER -->\n<div id="due-banner">\n  <div class="due-text">⏰ <span id="due-text"></span></div>\n  <button class="btn btn-orange btn-sm" onclick="showTab(\'fixed\')" data-i18n="view">View</button>\n</div>\n\n<!-- SEGMENTED TABS -->\n<div class="seg-wrap">\n  <div class="seg">\n    <button class="seg-item active" id="tt-overview" onclick="showTab(\'overview\')" data-i18n="tab_overview">Overview</button>\n    <button class="seg-item" id="tt-add" onclick="showTab(\'add\')" data-i18n="tab_add">Add</button>\n    <button class="seg-item" id="tt-history" onclick="showTab(\'history\')" data-i18n="tab_history">History</button>\n    <button class="seg-item" id="tt-fixed" onclick="showTab(\'fixed\')">\n      <span data-i18n="tab_fixed">Fixed</span><span class="tab-badge" id="fixed-badge" style="display:none"></span>\n    </button>\n    <button class="seg-item" id="tt-settings" onclick="showTab(\'settings\')" data-i18n="tab_settings">Settings</button>\n  </div>\n</div>\n\n<div class="main">\n\n  <!-- ── OVERVIEW ── -->\n  <div class="page active" id="page-overview">\n    <div class="acc-switcher">\n      <button class="acc-pill krw" id="at-krw" onclick="setAcc(\'krw\')">₩ Korean Won</button>\n      <button class="acc-pill" id="at-uyu" onclick="setAcc(\'uyu\')">$U Uruguayan Peso</button>\n      <button class="acc-pill" id="at-usd" onclick="setAcc(\'usd\')">US$ Dollar</button>\n    </div>\n    <div class="ov-grid" id="metrics"></div>\n    <div class="ov-charts">\n      <div class="card">\n        <div class="card-title" data-i18n="chart_flow">Money Flow</div>\n        <div style="position:relative;height:220px"><canvas id="barChart"></canvas></div>\n      </div>\n      <div class="card">\n        <div class="card-title" data-i18n="chart_cats">Spending by Category</div>\n        <div style="position:relative;height:220px"><canvas id="pieChart"></canvas></div>\n      </div>\n    </div>\n    <div class="card">\n      <div class="card-title" data-i18n="recent">Recent Transactions</div>\n      <div class="recent-preview">\n        <table><thead><tr>\n          <th data-i18n="col_date">Date</th>\n          <th data-i18n="col_desc">Description</th>\n          <th data-i18n="col_cat">Category</th>\n          <th>Account</th>\n          <th data-i18n="col_type">Type</th>\n          <th data-i18n="col_amount">Amount</th>\n          <th></th>\n        </tr></thead><tbody id="recent-body"></tbody></table>\n      </div>\n    </div>\n  </div>\n\n  <!-- ── ADD ── -->\n  <div class="page" id="page-add">\n    <div class="add-layout">\n      <div class="card">\n        <div class="type-toggle">\n          <button class="type-btn a-ex" id="btn-ex" onclick="setType(\'expense\')" data-i18n="btn_expense">Expense</button>\n          <button class="type-btn" id="btn-fund" onclick="setType(\'fund\')" data-i18n="btn_funds">Add funds</button>\n        </div>\n        <div class="form-grid">\n          <div class="field">\n            <label data-i18n="lbl_account">Account</label>\n            <select id="new-acc" onchange="updateSym()">\n              <option value="krw">₩ Korean Won</option>\n              <option value="uyu">$U Uruguayan Peso</option>\n              <option value="usd">US$ Dollar</option>\n            </select>\n          </div>\n          <div class="field">\n            <label data-i18n="lbl_amount">Amount</label>\n            <div class="amt-wrap">\n              <span class="amt-sym" id="amt-sym">₩</span>\n              <input type="number" id="new-amt" placeholder="0" step="1" min="0">\n            </div>\n          </div>\n          <div class="field">\n            <label data-i18n="lbl_date">Date</label>\n            <input type="date" id="new-date">\n          </div>\n          <div class="field" id="cat-group">\n            <label data-i18n="lbl_category">Category</label>\n            <select id="new-cat">\n              <option>Housing</option><option>Food</option><option>Transport</option>\n              <option>Entertainment</option><option>Health</option><option>Other</option>\n            </select>\n          </div>\n          <div class="field form-full">\n            <label data-i18n="lbl_description">Description</label>\n            <input type="text" id="new-desc" data-i18n-ph="ph_desc">\n          </div>\n        </div>\n        <button class="btn btn-full" id="submit-btn" onclick="submitEntry()" data-i18n="btn_save_expense">Save expense</button>\n        <p class="msg" id="add-msg"></p>\n      </div>\n      <div class="card">\n        <div class="card-title" data-i18n="recent">Recent</div>\n        <div class="recent-preview">\n          <table><thead><tr>\n            <th data-i18n="col_date">Date</th><th data-i18n="col_desc">Description</th>\n            <th>Acc</th><th data-i18n="col_amount">Amount</th><th></th>\n          </tr></thead><tbody id="add-recent-body"></tbody></table>\n        </div>\n      </div>\n    </div>\n  </div>\n\n  <!-- ── HISTORY ── -->\n  <div class="page" id="page-history">\n    <div class="card">\n      <div class="filter-bar">\n        <span class="sec-label" data-i18n="tab_history">Transaction History</span>\n        <div style="display:flex;gap:8px">\n          <select id="f-acc" onchange="loadHistory()">\n            <option value="" data-i18n="filter_all">All accounts</option>\n            <option value="krw">₩ Korean Won</option>\n            <option value="uyu">$U Pesos</option>\n            <option value="usd">US$ Dollar</option>\n          </select>\n          <input type="month" id="f-month" onchange="loadHistory()">\n        </div>\n      </div>\n      <table><thead><tr>\n        <th data-i18n="col_date">Date</th>\n        <th data-i18n="col_desc">Description</th>\n        <th data-i18n="col_cat">Category</th>\n        <th>Account</th>\n        <th data-i18n="col_type">Type</th>\n        <th data-i18n="col_amount">Amount</th>\n        <th></th>\n      </tr></thead><tbody id="hist-body"></tbody></table>\n    </div>\n  </div>\n\n  <!-- ── FIXED PAYMENTS ── -->\n  <div class="page" id="page-fixed">\n    <div class="fixed-layout">\n      <div>\n        <div class="card fixed-list-card" id="fp-due-section" style="display:none">\n          <div class="card-title" data-i18n="fp_due_now">Due This Month</div>\n          <div class="fp-grid" id="fp-due-list"></div>\n        </div>\n        <div class="card fixed-list-card">\n          <div class="card-title" data-i18n="fp_all">All Fixed Payments</div>\n          <div class="fp-grid" id="fp-all-list"></div>\n        </div>\n      </div>\n      <div class="card" style="position:sticky;top:86px">\n        <div class="card-title" data-i18n="fp_new">New Fixed Payment</div>\n        <div style="display:flex;flex-direction:column;gap:10px">\n          <div class="field">\n            <label data-i18n="fp_name">Name</label>\n            <input type="text" id="fp-name" data-i18n-ph="fp_name_ph">\n          </div>\n          <div style="display:grid;grid-template-columns:1fr 1fr;gap:10px">\n            <div class="field">\n              <label data-i18n="lbl_account">Account</label>\n              <select id="fp-acc" onchange="updateFpSym()">\n                <option value="krw">₩ Korean Won</option>\n                <option value="uyu">$U Pesos</option>\n                <option value="usd">US$ Dollar</option>\n              </select>\n            </div>\n            <div class="field">\n              <label data-i18n="lbl_amount">Amount</label>\n              <div class="amt-wrap">\n                <span class="amt-sym" id="fp-sym">₩</span>\n                <input type="number" id="fp-amt" placeholder="0" step="1" min="0">\n              </div>\n            </div>\n          </div>\n          <div style="display:grid;grid-template-columns:1fr 1fr;gap:10px">\n            <div class="field">\n              <label data-i18n="fp_day">Day of month</label>\n              <input type="number" id="fp-day" placeholder="1–31" min="1" max="31">\n            </div>\n            <div class="field">\n              <label data-i18n="lbl_category">Category</label>\n              <select id="fp-cat">\n                <option>Housing</option><option>Food</option><option>Transport</option>\n                <option>Entertainment</option><option>Health</option><option>Other</option>\n              </select>\n            </div>\n          </div>\n        </div>\n        <div class="mt12">\n          <button class="btn btn-blue btn-full" onclick="addFixed()" data-i18n="fp_save">Save fixed payment</button>\n        </div>\n        <p class="msg" id="fp-msg"></p>\n      </div>\n    </div>\n  </div>\n\n  <!-- ── SETTINGS ── -->\n  <div class="page" id="page-settings">\n    <p style="font-size:13px;color:var(--lbl3);margin-bottom:18px" data-i18n="settings_desc">Set these to your real bank balances. This corrects the number without creating a transaction.</p>\n    <div class="settings-layout">\n      <div class="card">\n        <div class="card-title" style="color:var(--blue)">₩ Korean Won (KRW)</div>\n        <div class="settings-row">\n          <input type="number" id="bal-krw" placeholder="0" step="1">\n          <button class="btn btn-blue btn-sm" onclick="saveBalance(\'krw\')" data-i18n="btn_save">Save</button>\n        </div>\n        <p class="msg" id="msg-krw"></p>\n      </div>\n      <div class="card">\n        <div class="card-title" style="color:var(--orange)">$U Uruguayan Peso (UYU)</div>\n        <div class="settings-row">\n          <input type="number" id="bal-uyu" placeholder="0.00" step="0.01">\n          <button class="btn btn-orange btn-sm" onclick="saveBalance(\'uyu\')" data-i18n="btn_save">Save</button>\n        </div>\n        <p class="msg" id="msg-uyu"></p>\n      </div>\n      <div class="card">\n        <div class="card-title" style="color:var(--green)">US$ Dollar (USD)</div>\n        <div class="settings-row">\n          <input type="number" id="bal-usd" placeholder="0.00" step="0.01">\n          <button class="btn btn-green btn-sm" onclick="saveBalance(\'usd\')" data-i18n="btn_save">Save</button>\n        </div>\n        <p class="msg" id="msg-usd"></p>\n      </div>\n    </div>\n  </div>\n\n</div><!-- .main -->\n</div><!-- #app -->\n\n<script>\n/* ── I18N ── */\nconst STR={\n  en:{\n    tab_overview:"Overview",tab_add:"Add",tab_history:"History",\n    tab_fixed:"Fixed",tab_settings:"Settings",\n    chart_flow:"Money Flow",chart_cats:"Spending by Category",recent:"Recent Transactions",\n    col_date:"Date",col_desc:"Description",col_cat:"Category",col_type:"Type",col_amount:"Amount",\n    btn_expense:"Expense",btn_funds:"Add funds",\n    lbl_account:"Account",lbl_amount:"Amount",lbl_date:"Date",lbl_category:"Category",lbl_description:"Description",\n    ph_desc:"e.g. Supermarket",\n    btn_save_expense:"Save expense",btn_save_funds:"Add funds",btn_save:"Save",\n    filter_all:"All accounts",\n    settings_desc:"Set these to your real bank balances. This corrects the number without creating a transaction.",\n    metric_bal:"Balance",metric_spent:"Spent (30 days)",metric_txns:"Total transactions",\n    metric_current:"current balance",metric_all:"across all accounts",\n    saved_ok:"Saved — new balance:",bal_updated:"Balance updated to",\n    fill_fields:"Please fill in amount and description.",del_confirm:"Delete this transaction? Balance will be reversed.",\n    no_txns:"No transactions yet",no_txns_period:"No transactions for this period",\n    lbl_funds:"funds",lbl_expense:"expense",view:"View",\n    fp_due_now:"Due This Month",fp_all:"All Fixed Payments",fp_new:"New Fixed Payment",\n    fp_name:"Name",fp_name_ph:"e.g. Rent",fp_day:"Day of month",fp_save:"Save fixed payment",\n    fp_apply:"Apply now",fp_undo:"Undo",fp_paid:"✓ Paid",fp_del_confirm:"Delete this fixed payment?",\n    fp_none:"No fixed payments yet",fp_due_banner:"fixed payment(s) due this month",\n  },\n  es:{\n    tab_overview:"Resumen",tab_add:"Agregar",tab_history:"Historial",\n    tab_fixed:"Fijos",tab_settings:"Ajustes",\n    chart_flow:"Flujo de dinero",chart_cats:"Gastos por categoría",recent:"Transacciones recientes",\n    col_date:"Fecha",col_desc:"Descripción",col_cat:"Categoría",col_type:"Tipo",col_amount:"Monto",\n    btn_expense:"Gasto",btn_funds:"Agregar fondos",\n    lbl_account:"Cuenta",lbl_amount:"Monto",lbl_date:"Fecha",lbl_category:"Categoría",lbl_description:"Descripción",\n    ph_desc:"ej. Supermercado",\n    btn_save_expense:"Guardar gasto",btn_save_funds:"Agregar fondos",btn_save:"Guardar",\n    filter_all:"Todas las cuentas",\n    settings_desc:"Ingresá el saldo real de tu banco. Corrige el número sin crear una transacción.",\n    metric_bal:"Saldo",metric_spent:"Gastado (30 días)",metric_txns:"Total transacciones",\n    metric_current:"saldo actual",metric_all:"todas las cuentas",\n    saved_ok:"Guardado — nuevo saldo:",bal_updated:"Saldo actualizado a",\n    fill_fields:"Completá el monto y la descripción.",del_confirm:"¿Eliminar? El saldo se revertirá.",\n    no_txns:"Sin transacciones aún",no_txns_period:"Sin transacciones en este período",\n    lbl_funds:"fondos",lbl_expense:"gasto",view:"Ver",\n    fp_due_now:"Vencen este mes",fp_all:"Todos los pagos fijos",fp_new:"Nuevo pago fijo",\n    fp_name:"Nombre",fp_name_ph:"ej. Alquiler",fp_day:"Día del mes",fp_save:"Guardar pago fijo",\n    fp_apply:"Aplicar ahora",fp_undo:"Deshacer",fp_paid:"✓ Pagado",fp_del_confirm:"¿Eliminar este pago fijo?",\n    fp_none:"Sin pagos fijos aún",fp_due_banner:"pago(s) fijo(s) pendiente(s) este mes",\n  }\n};\nlet lang=localStorage.getItem(\'clarifi_lang\')||\'en\';\nfunction T(k){return(STR[lang]||STR.en)[k]||k}\nfunction applyLang(){\n  document.documentElement.lang=lang;\n  document.getElementById(\'lang-btn\').textContent=\'🌐 \'+(lang===\'en\'?\'ES\':\'EN\');\n  document.querySelectorAll(\'[data-i18n]\').forEach(el=>el.textContent=T(el.dataset.i18n));\n  document.querySelectorAll(\'[data-i18n-ph]\').forEach(el=>el.placeholder=T(el.dataset.i18nPh));\n  setType(curType);\n  if(loaded){renderAll();renderFixed();}\n}\nfunction toggleLang(){lang=lang===\'en\'?\'es\':\'en\';localStorage.setItem(\'clarifi_lang\',lang);applyLang();}\n\n/* ── THEME ── */\nlet theme=localStorage.getItem(\'clarifi_theme\')||\'dark\';\nfunction applyTheme(){\n  document.documentElement.setAttribute(\'data-theme\',theme);\n  document.getElementById(\'theme-btn\').textContent=theme===\'dark\'?\'☀️ Light\':\'🌙 Dark\';\n  if(loaded) renderCharts();\n}\nfunction toggleTheme(){theme=theme===\'dark\'?\'light\':\'dark\';localStorage.setItem(\'clarifi_theme\',theme);applyTheme();}\n\n/* ── CONSTANTS ── */\nconst ACCOUNTS={\n  krw:{symbol:\'₩\',decimals:0},\n  uyu:{symbol:\'$U\',decimals:2},\n  usd:{symbol:\'US$\',decimals:2},\n};\nconst SYM={krw:\'₩\',uyu:\'$U\',usd:\'US$\'};\nconst PAL=[\'#0A84FF\',\'#FF9F0A\',\'#30D158\',\'#FF375F\',\'#BF5AF2\',\'#40CBE0\',\'#FFD60A\'];\n\n/* ── FORMAT ── */\nfunction todayStr(){return new Date().toISOString().slice(0,10);}\nfunction roundAcc(acc,v){const d=(ACCOUNTS[acc]||{decimals:2}).decimals;return Math.round(Number(v)*10**d)/10**d;}\nfunction fmtFull(n,acc){\n  const d=(ACCOUNTS[acc]||{decimals:2}).decimals;\n  const abs=Math.abs(Number(n));\n  const s=abs.toLocaleString(\'en-US\',{minimumFractionDigits:d,maximumFractionDigits:d});\n  const p=s.split(\'.\');\n  return{sym:SYM[acc]||\'\',int:p[0],dec:p[1]?\'.\'+p[1]:\'\',neg:Number(n)<0};\n}\nfunction fmtStr(n,acc){const f=fmtFull(n,acc);return(f.neg?\'− \':\'\')+f.sym+\' \'+f.int+f.dec;}\nfunction metricHTML(n,acc){\n  const f=fmtFull(n,acc);\n  const cls=n<0?\'v-red\':acc===\'krw\'?\'v-krw\':acc===\'uyu\'?\'v-uyu\':acc===\'usd\'?\'v-usd\':\'\';\n  return`<div class="metric-val ${cls}">${f.neg?\'<span style="opacity:.45">−</span>\':\'\'}<span class="sym">${f.sym}</span>${f.int}<span class="dec">${f.dec}</span></div>`;\n}\nfunction accChip(acc){return`<span class="chip chip-${acc}">${SYM[acc]}</span>`;}\n\n/* ── STATE ── */\nlet curAcc=\'krw\',curType=\'expense\',loaded=false,summary={};\nlet barChart,pieChart;\n\n/* ── TABS ── */\nconst TABS=[\'overview\',\'add\',\'history\',\'fixed\',\'settings\'];\nfunction showTab(name){\n  if(!TABS.includes(name)) return;\n  const wasOverview=document.getElementById(\'page-overview\').classList.contains(\'active\');\n  TABS.forEach(t=>{\n    document.getElementById(\'page-\'+t).classList.toggle(\'active\',t===name);\n    document.getElementById(\'tt-\'+t).classList.toggle(\'active\',t===name);\n  });\n  if(name===\'overview\'&&!wasOverview) renderCharts();\n  if(name===\'history\') loadHistory();\n  if(name===\'fixed\') renderFixed();\n  if(name===\'settings\'){\n    const b=summary.balances||{};\n    document.getElementById(\'bal-krw\').value=Math.round(b.krw||0);\n    document.getElementById(\'bal-uyu\').value=(b.uyu||0).toFixed(2);\n    document.getElementById(\'bal-usd\').value=(b.usd||0).toFixed(2);\n  }\n}\n\n/* ── ACCOUNT + TYPE ── */\nfunction setAcc(acc){\n  curAcc=acc;\n  [\'krw\',\'uyu\',\'usd\'].forEach(a=>document.getElementById(\'at-\'+a).className=\'acc-pill\'+(a===acc?\' \'+a:\'\'));\n  renderMetrics(); renderCharts();\n}\nfunction setType(type){\n  curType=type;\n  document.getElementById(\'btn-ex\').className=\'type-btn\'+(type===\'expense\'?\' a-ex\':\'\');\n  document.getElementById(\'btn-fund\').className=\'type-btn\'+(type===\'fund\'?\' a-fund\':\'\');\n  document.getElementById(\'cat-group\').style.display=type===\'expense\'?\'\':\'none\';\n  document.getElementById(\'submit-btn\').textContent=type===\'expense\'?T(\'btn_save_expense\'):T(\'btn_save_funds\');\n  document.getElementById(\'submit-btn\').className=\'btn btn-full btn-\'+(type===\'expense\'?\'blue\':\'green\');\n}\nfunction updateSym(){\n  const acc=document.getElementById(\'new-acc\').value;\n  document.getElementById(\'amt-sym\').textContent=SYM[acc];\n  document.getElementById(\'new-amt\').step=(ACCOUNTS[acc]||{decimals:2}).decimals===0?\'1\':\'0.01\';\n}\nfunction updateFpSym(){\n  document.getElementById(\'fp-sym\').textContent=SYM[document.getElementById(\'fp-acc\').value];\n}\n\n/* ── DATA ── */\nasync function apiFetch(url,opts={}){\n  const r=await fetch(url,{headers:{\'Content-Type\':\'application/json\'},...opts});\n  return r.json();\n}\n\nasync function renderAll(){\n  summary=await apiFetch(\'/api/summary\');\n  const b=summary.balances||{};\n  document.getElementById(\'nav-krw\').textContent=fmtStr(b.krw||0,\'krw\');\n  document.getElementById(\'nav-uyu\').textContent=fmtStr(b.uyu||0,\'uyu\');\n  document.getElementById(\'nav-usd\').textContent=fmtStr(b.usd||0,\'usd\');\n  renderMetrics(); renderCharts(); renderRecent(); checkDueBanner();\n  renderAddRecent();\n}\n\nfunction renderMetrics(){\n  if(!summary.balances) return;\n  const b=summary.balances[curAcc]||0, st=(summary.stats||{})[curAcc]||{};\n  document.getElementById(\'metrics\').innerHTML=`\n    <div class="card metric">\n      <div class="metric-ey">${T(\'metric_bal\')}</div>\n      ${metricHTML(b,curAcc)}\n      <div class="metric-sub">${T(\'metric_current\')}</div>\n    </div>\n    <div class="card metric">\n      <div class="metric-ey">${T(\'metric_spent\')}</div>\n      ${metricHTML(st.last30||0,curAcc)}\n    </div>\n    <div class="card metric">\n      <div class="metric-ey">${T(\'metric_txns\')}</div>\n      <div class="metric-val" style="font-weight:300">${summary.total_txns||0}</div>\n      <div class="metric-sub">${T(\'metric_all\')}</div>\n    </div>`;\n}\n\nfunction cc(){return{\n  g:theme===\'dark\'?\'rgba(255,255,255,.05)\':\'rgba(60,60,80,.06)\',\n  t:theme===\'dark\'?\'rgba(240,240,248,.40)\':\'rgba(13,13,24,.42)\'\n};}\n\nfunction renderCharts(){\n  if(!summary.stats||!document.getElementById(\'page-overview\').classList.contains(\'active\')) return;\n  const st=(summary.stats||{})[curAcc]||{},monthly=st.monthly||{},months=Object.keys(monthly),c=cc();\n  const ci=curAcc===\'krw\'?\'#0A84FF\':curAcc===\'uyu\'?\'#FF9F0A\':\'#30D158\';\n  const cif=curAcc===\'krw\'?\'rgba(10,132,255,.55)\':curAcc===\'uyu\'?\'rgba(255,159,10,.50)\':\'rgba(48,209,88,.55)\';\n  if(barChart) barChart.destroy();\n  if(months.length){\n    const labels=months.map(m=>{const[y,mo]=m.split(\'-\');return new Date(y,mo-1).toLocaleString(\'default\',{month:\'short\',year:\'2-digit\'});});\n    barChart=new Chart(document.getElementById(\'barChart\'),{type:\'bar\',data:{labels,datasets:[\n      {label:lang===\'en\'?\'Funds in\':\'Entrada\',data:months.map(m=>monthly[m].in),backgroundColor:cif,borderColor:ci,borderWidth:1,borderRadius:6},\n      {label:lang===\'en\'?\'Expenses\':\'Gastos\',data:months.map(m=>monthly[m].out),backgroundColor:\'rgba(255,69,58,.45)\',borderColor:\'#FF453A\',borderWidth:1,borderRadius:6}\n    ]},options:{responsive:true,maintainAspectRatio:false,animation:false,resizeDelay:120,\n      plugins:{legend:{labels:{font:{size:12,family:\'Inter\'},color:c.t,boxWidth:10,padding:14}}},\n      scales:{x:{grid:{display:false},ticks:{font:{size:11},color:c.t}},\n              y:{grid:{color:c.g},border:{dash:[4,4]},ticks:{font:{size:11},color:c.t,callback:v=>SYM[curAcc]+\' \'+v.toLocaleString()}}}\n    }});\n  }\n  const ec=st.exp_cat||{},cats=Object.keys(ec);\n  if(pieChart) pieChart.destroy();\n  if(cats.length){\n    pieChart=new Chart(document.getElementById(\'pieChart\'),{type:\'doughnut\',data:{labels:cats,datasets:[{data:cats.map(cat=>ec[cat]),backgroundColor:PAL,borderWidth:0,hoverOffset:6}]},\n      options:{responsive:true,maintainAspectRatio:false,animation:false,resizeDelay:120,cutout:\'63%\',\n        plugins:{legend:{position:\'right\',labels:{font:{size:12,family:\'Inter\'},color:c.t,boxWidth:10,padding:12}}}}});\n  }\n}\n\nfunction txnRow(t,cols=7,showDel=true){\n  const isFund=t.type===\'fund\',acc=t.account||\'krw\';\n  const catCell=cols===7?`<td style="color:var(--lbl2)">${t.category||\'\'}</td>`:\'\';\n  return`<tr>\n    <td style="color:var(--lbl2);white-space:nowrap">${t.date||\'\'}</td>\n    <td style="font-weight:500">${t.description||\'\'}</td>\n    ${catCell}\n    <td>${accChip(acc)}</td>\n    <td><span class="chip chip-${t.type}">${isFund?T(\'lbl_funds\'):T(\'lbl_expense\')}</span></td>\n    <td class="td-mono" style="color:${isFund?\'var(--green)\':\'var(--red)\'}">\n      ${isFund?\'+\':\'−\'} ${fmtStr(t.amount,acc)}</td>\n    <td>${showDel?`<button class="del-btn" onclick="delTxn(${t.id})" title="Delete">×</button>`:\'\'}</td>\n  </tr>`;\n}\nfunction renderRecent(){\n  const rows=(summary.recent||[]).map(t=>txnRow(t,7)).join(\'\');\n  document.getElementById(\'recent-body\').innerHTML=rows||`<tr><td colspan="7" class="empty">${T(\'no_txns\')}</td></tr>`;\n}\nfunction renderAddRecent(){\n  const rows=(summary.recent||[]).slice(0,8).map(t=>txnRow(t,5)).join(\'\');\n  document.getElementById(\'add-recent-body\').innerHTML=rows||`<tr><td colspan="5" class="empty">${T(\'no_txns\')}</td></tr>`;\n}\nasync function loadHistory(){\n  const fm=document.getElementById(\'f-month\').value,fa=document.getElementById(\'f-acc\').value;\n  let txns=await apiFetch(\'/api/transactions\');\n  if(fm) txns=txns.filter(t=>t.date&&t.date.startsWith(fm));\n  if(fa) txns=txns.filter(t=>(t.account||\'krw\')===fa);\n  txns.sort((a,b)=>b.date.localeCompare(a.date));\n  document.getElementById(\'hist-body\').innerHTML=txns.map(t=>txnRow(t,7)).join(\'\')||`<tr><td colspan="7" class="empty">${T(\'no_txns_period\')}</td></tr>`;\n}\n\nasync function submitEntry(){\n  const amt=parseFloat(document.getElementById(\'new-amt\').value);\n  const desc=document.getElementById(\'new-desc\').value.trim();\n  const acc=document.getElementById(\'new-acc\').value;\n  const msg=document.getElementById(\'add-msg\');\n  if(!amt||!desc){msg.textContent=T(\'fill_fields\');msg.className=\'msg err\';return;}\n  const payload={amount:amt,description:desc,account:acc,\n    date:document.getElementById(\'new-date\').value||todayStr(),\n    category:document.getElementById(\'new-cat\').value};\n  const d=await apiFetch(\'/api/\'+(curType===\'fund\'?\'fund\':\'expense\'),{method:\'POST\',body:JSON.stringify(payload)});\n  msg.textContent=T(\'saved_ok\')+\' \'+fmtStr(d.balance,acc);\n  msg.className=\'msg ok\';\n  document.getElementById(\'new-amt\').value=\'\';\n  document.getElementById(\'new-desc\').value=\'\';\n  renderAll();\n  setTimeout(()=>{msg.textContent=\'\';},4000);\n}\nasync function delTxn(id){\n  if(!confirm(T(\'del_confirm\'))) return;\n  await apiFetch(\'/api/transactions/\'+id,{method:\'DELETE\'});\n  renderAll();\n  if(document.getElementById(\'page-history\').classList.contains(\'active\')) loadHistory();\n}\nasync function saveBalance(acc){\n  const val=parseFloat(document.getElementById(\'bal-\'+acc).value);\n  const msg=document.getElementById(\'msg-\'+acc);\n  if(isNaN(val)) return;\n  await apiFetch(\'/api/balance\',{method:\'POST\',body:JSON.stringify({account:acc,balance:val})});\n  msg.textContent=T(\'bal_updated\')+\' \'+fmtStr(val,acc);\n  msg.className=\'msg ok\';\n  renderAll();\n  setTimeout(()=>{msg.textContent=\'\';},4000);\n}\n\n/* ── FIXED PAYMENTS ── */\nfunction checkDueBanner(){\n  const count=summary.due_count||0;\n  const badge=document.getElementById(\'fixed-badge\');\n  const banner=document.getElementById(\'due-banner\');\n  if(count>0){\n    badge.style.display=\'inline-block\';badge.textContent=count;\n    banner.style.display=\'flex\';\n    document.getElementById(\'due-text\').textContent=count+\' \'+T(\'fp_due_banner\');\n  } else {\n    badge.style.display=\'none\';banner.style.display=\'none\';\n  }\n}\n\nfunction fpHTML(fp){\n  const acc=fp.account||\'krw\',due=fp.due_this_month,applied=fp.applied_this_month;\n  const mo=new Date().toLocaleString(\'default\',{month:\'short\'});\n  const ord=n=>n+(n==1?\'st\':n==2?\'nd\':n==3?\'rd\':\'th\');\n  return`<div class="fp-row${due?\' due\':\'\'}">\n    <div class="fp-badge ${acc}">\n      <span class="fp-d">${fp.day}</span>\n      <span class="fp-mo">${mo}</span>\n    </div>\n    <div class="fp-info">\n      <div class="fp-name">${fp.name}</div>\n      <div class="fp-meta">${fp.category} · every ${lang===\'en\'?ord(fp.day):\'día \'+fp.day}</div>\n    </div>\n    <div class="fp-right">\n      <div class="fp-amt" style="color:${acc===\'krw\'?\'var(--blue)\':acc===\'uyu\'?\'var(--orange)\':\'var(--green)\'}">${fmtStr(fp.amount,acc)}</div>\n      <div class="fp-actions">\n        ${due?`<button class="btn btn-orange btn-sm" onclick="applyFixed(${fp.id})">${T(\'fp_apply\')}</button>`:\'\'}\n        ${applied?`<span class="chip chip-ok">${T(\'fp_paid\')}</span>\n          <button class="btn btn-ghost btn-sm" onclick="undoFixed(${fp.id})">${T(\'fp_undo\')}</button>`:\'\'}\n        <button class="del-btn" onclick="delFixed(${fp.id})">×</button>\n      </div>\n    </div>\n  </div>`;\n}\n\nasync function renderFixed(){\n  const fixed=await apiFetch(\'/api/fixed\');\n  const due=fixed.filter(f=>f.due_this_month);\n  const dueSec=document.getElementById(\'fp-due-section\');\n  dueSec.style.display=due.length?\'block\':\'none\';\n  document.getElementById(\'fp-due-list\').innerHTML=due.map(fpHTML).join(\'\');\n  document.getElementById(\'fp-all-list\').innerHTML=fixed.length?fixed.map(fpHTML).join(\'\'):`<p class="empty">${T(\'fp_none\')}</p>`;\n}\nasync function addFixed(){\n  const name=document.getElementById(\'fp-name\').value.trim();\n  const amt=parseFloat(document.getElementById(\'fp-amt\').value);\n  const acc=document.getElementById(\'fp-acc\').value;\n  const day=parseInt(document.getElementById(\'fp-day\').value);\n  const cat=document.getElementById(\'fp-cat\').value;\n  const msg=document.getElementById(\'fp-msg\');\n  if(!name||!amt||!day||day<1||day>31){msg.textContent=T(\'fill_fields\');msg.className=\'msg err\';return;}\n  await apiFetch(\'/api/fixed\',{method:\'POST\',body:JSON.stringify({name,amount:amt,account:acc,category:cat,day})});\n  document.getElementById(\'fp-name\').value=\'\';\n  document.getElementById(\'fp-amt\').value=\'\';\n  document.getElementById(\'fp-day\').value=\'\';\n  msg.textContent=\'✓ Saved!\';msg.className=\'msg ok\';\n  setTimeout(()=>{msg.textContent=\'\';},3000);\n  renderFixed();renderAll();\n}\nasync function delFixed(id){\n  if(!confirm(T(\'fp_del_confirm\'))) return;\n  await apiFetch(\'/api/fixed/\'+id,{method:\'DELETE\'});\n  renderFixed();renderAll();\n}\nasync function applyFixed(id){\n  await apiFetch(\'/api/fixed/\'+id+\'/apply\',{method:\'POST\'});\n  renderAll();renderFixed();\n}\nasync function undoFixed(id){\n  await apiFetch(\'/api/fixed/\'+id+\'/undo\',{method:\'POST\'});\n  renderAll();renderFixed();\n}\n\n/* ── INIT ── */\nconst today=todayStr();\ndocument.getElementById(\'new-date\').value=today;\ndocument.getElementById(\'f-month\').value=today.slice(0,7);\nsetType(\'expense\');\napplyTheme();\napplyLang();\nloaded=true;\nrenderAll().finally(()=>{document.getElementById(\'app\').dataset.ready=\'true\';});\n</script>\n</body>\n</html>\n'


if __name__ == '__main__':
    init_data()
    port = int(os.environ.get('PORT', 5000))
    print('\n  ClariFi Dashboard -> http://localhost:' + str(port) + '\n')
    app.run(host='127.0.0.1', port=port, debug=False)
